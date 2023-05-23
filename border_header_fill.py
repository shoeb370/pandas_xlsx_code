# -*- coding: utf-8 -*-
"""
Created on Tue May 23 14:24:45 2023

@author: Shoeb
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font




def concat_final_df(api_df,outbond):
    final_df = pd.concat([outbond, api_df], ignore_index=True)
    
    
    vc_Project = final_df.Project.value_counts()

    col1=[i for i in vc_Project.index]
    col2=[i for i in vc_Project]
    total = col2[0] + col2[1]
    df = pd.DataFrame(list(zip(col1, col2)), columns = ['Column1 Name', 'Counts'])
    
    new_row = pd.DataFrame({'Column1 Name':['Total'],'Counts':[total]})
    # df = df.append(new_row, ignore_index=True)
    
    df =pd.concat([df,new_row],ignore_index=True)
    
    value_counts_Funder = final_df.FunderName.value_counts()

    col3=[i for i in value_counts_Funder.index]
    col4=[i for i in value_counts_Funder]
    total1 = sum(col4)
    df2 = pd.DataFrame(list(zip(col3, col4)), columns = ['Column2 Name', 'Counts'])  
    new_row1 =pd.DataFrame({'Column2 Name':['Total'],'Counts':[total1]})
    df2=pd.concat([df2,new_row1],ignore_index=True)
    # df2 = df2.append(new_row, ignore_index=True)
    
    
    with pd.ExcelWriter(r".\Final_file.xlsx") as writer:
        final_df.to_excel(writer,sheet_name="Sheet 1 name",index=False)
        df.to_excel(writer,sheet_name="Sheet 2 name",index=False)    
        df2.to_excel(writer,sheet_name="Sheet 3 name",index=False)
        
    return


def border():
    workbook = load_workbook('Final_file.xlsx')
    for sheet in workbook.sheetnames:
        # Select the current sheet
        current_sheet = workbook[sheet]

        # Get the dimensions (dynamic range) of the sheet
        dimensions = current_sheet.dimensions

        # Create the border object
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Iterate over the cells in the dynamic range and apply the border style
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        for row in current_sheet[dimensions]:
            for cell in row:
                cell.border = border
                # Check if it's a header cell (first row)
                if cell.row == 1:
                # Apply fill color
                    cell.fill = fill #to fill the headers color

                # Apply font style (optional)
                    cell.font = Font(bold=True)
                    
    workbook.save('Final_file.xlsx')

api_df = pd.read_csv(r"./file_track_1.csv")
outbond = pd.read_csv(r"./file_track_2.csv")



concat_final_df(api_df, outbond)
border()
